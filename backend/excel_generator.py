from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from typing import Dict, List
import os

class ExcelGenerator:
    """Generate Excel files from processed data"""

    def __init__(self, options: Dict = None):
        self.options = options or {}
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Data'

    def generate(self, data_list: List[Dict], output_path: str) -> str:
        """
        Generate Excel file
        
        Args:
            data_list: List of processed data dictionaries
            output_path: Output file path
            
        Returns:
            Path to generated file
        """
        # Write data to worksheet
        self._write_data(data_list)
        
        # Apply formatting
        self._apply_formatting()
        
        # Create charts if requested
        if self.options.get('createChart', False):
            self._create_charts(data_list)
        
        # Save file
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        self.wb.save(output_path)
        
        return output_path

    def _write_data(self, data_list: List[Dict]):
        """Write data to worksheet"""
        row = 1
        
        for data_idx, data in enumerate(data_list):
            # Write headers
            headers = data.get('headers', [])
            for col, header in enumerate(headers, 1):
                self.ws.cell(row=row, column=col, value=header)
            
            row += 1
            
            # Write data rows
            rows = data.get('rows', [])
            for data_row in rows:
                for col, header in enumerate(headers, 1):
                    value = data_row.get(header, '')
                    self.ws.cell(row=row, column=col, value=value)
                row += 1
            
            # Add separator between datasets
            if data_idx < len(data_list) - 1:
                row += 1

    def _apply_formatting(self):
        """Apply Excel formatting"""
        # Header styling
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header formatting
        for col in range(1, self.ws.max_column + 1):
            cell = self.ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Apply border to all cells
        for row in self.ws.iter_rows(min_row=1, max_row=self.ws.max_row, min_col=1, max_col=self.ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Auto-adjust column width
        for col in range(1, self.ws.max_column + 1):
            max_length = 0
            column_letter = self.ws.cell(row=1, column=col).column_letter
            
            for cell in self.ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width
        
        # Auto-adjust row height
        self.ws.row_dimensions[1].height = 25

    def _create_charts(self, data_list: List[Dict]):
        """Create charts from data"""
        if not data_list or not data_list[0].get('rows'):
            return
        
        data = data_list[0]
        rows = data.get('rows', [])
        headers = data.get('headers', [])
        
        if len(rows) < 2 or len(headers) < 2:
            return
        
        try:
            # Create a bar chart
            chart = BarChart()
            chart.title = 'Data Chart'
            chart.y_axis.title = headers[1] if len(headers) > 1 else 'Value'
            chart.x_axis.title = headers[0]
            
            # Add data to chart
            data_range = Reference(self.ws, min_col=2, min_row=1, max_row=len(rows) + 1, max_col=2)
            cats = Reference(self.ws, min_col=1, min_row=2, max_row=len(rows) + 1)
            chart.add_data(data_range, titles_from_data=True)
            chart.set_categories(cats)
            
            # Add chart to worksheet
            chart_ws = self.wb.create_sheet('Charts')
            chart_ws.add_chart(chart, "A1")
        except Exception as e:
            print(f"Error creating chart: {e}")
